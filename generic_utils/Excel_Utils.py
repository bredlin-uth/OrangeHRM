import os

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from generic_utils import Config_Utils, Common_Utils

excel_path = os.path.join(os.path.dirname(os.path.abspath('.')), "test_data\\Data.xlsx")


def get_row_count(sheet_name):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(sheet_name):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data_from_excel(sheet_name, row, column):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row, column).value


def write_data_into_excel(sheet_name, row, column, data):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    sheet.cell(row, column).value = data
    workbook.save(excel_path)


def fetch_data_from_excel(sheet_name, row_header, column_header):
    workbook = openpyxl.open(excel_path)
    sheet = workbook[sheet_name]
    # Find the row and column indices
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_header in row:
            break
    for col_index, col in enumerate(sheet.iter_cols(values_only=True), start=1):
        if column_header in col:
            break
    # Check if the headers were found
    if row_index and col_index:
        cell_value = sheet.cell(row_index, col_index).value
        return cell_value
    else:
        return None


def get_row_excel_data(sheet_name, row, path=excel_path):
    data = []
    workbook = openpyxl.open(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column
    for i in range(row, total_rows + 1):
        for column in range(1, total_columns + 1):
            data.append(sheet.cell(i, column).value)
    return data


def fetch_data_by_row_header(sheet_name, row_header_value, path=excel_path):
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]

        # Find the row index based on the row header value
        row_index = None
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_header_value in row:
                row_index = row_num
                break
        print(row_index)
        if row_index is None:
            raise ValueError(f"Row header '{row_header_value}' not found in sheet '{sheet_name}'.")

        # Extract data from the specified row
        data = []
        for cell in sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True):
            print()
            data.append(list(cell))

        return data
    except Exception as e:
        print(f"Error: {e}")
        return []


def get_row_number(row_header, sheet_name, path=excel_path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    row_index = None
    for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_header in row:
            row_index = row_num
            break
    print(row_index)


def get_column_number(column_header, sheet_name, path=excel_path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    col_index = None
    for col_index, col in enumerate(sheet.iter_cols(values_only=True), start=1):
        if column_header in col:
            break
    print(col_index)


def extract_whole_data_as_dicts(filename, sheet_name):
    """
    Extracts data from an Excel sheet into a list of dictionaries, where each dictionary represents a row and keys are column headers.

    Args:
        filename: The name of the Excel file.
        sheet_name: The name of the worksheet within the file.

    Returns:
        A list of dictionaries, where each dictionary represents a row.
    """

    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]

        # Extract column headers
        headers = [cell.value for cell in sheet[1]]

        # Extract data rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        return data

    except Exception as e:
        print(f"Error: {e}")
        return []


def extract_data_as_dicts(row_header, sheet_name, filename=excel_path):
    """
    Extracts the row data from an Excel sheet into a dictionary (column header, data) for the specified row header.

    Args:
        row_header: The value of the row header from which to start extracting data.
        sheet_name: The name of the worksheet within the file.
        filename: The name of the Excel file.

    Returns:
        A list of dictionaries, where each dictionary represents a row.
    """

    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]

        # Find the row index based on the row header value
        row_index = None
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_header in row:
                row_index = row_num
                break

        if row_index is None:
            raise ValueError(f"Row header '{row_header}' not found in sheet '{sheet_name}'.")

        # Extract column headers
        headers = [cell.value for cell in sheet[1]]

        # Extract data rows
        row_dict = ()
        for row in sheet.iter_rows(min_row=row_index, values_only=True):
            print(row)
            row_dict = dict(zip(headers, row))
            break
        return row_dict

    except Exception as e:
        print(f"Error: {e}")
        return ()


def fetch_data_as_dicts(sheet_name, filename=excel_path):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
        return row_dict
    except Exception as e:
        print(f"Error: {e}")
        return None


def create_a_excel_file(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(file_name)
    return file_name


def create_sheet_in_excel(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb.create_sheet(title=sheet_name)
    wb.save(file_path)

def write_data_to_excel(file_path, sheet_name, data):
    # Load the workbook if it exists, otherwise create a new one
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()
    # Get the sheet by name, create if it doesn't exist
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    # Write data to the sheet
    last_row = ws.max_row + 1
    if isinstance(data, (list, tuple)):
        for idx, val in enumerate(data, start=1):
            print(val)
            ws.cell(row=last_row, column=idx).value = val
    elif isinstance(data, dict):
        for idx, (key, val) in enumerate(data.items(), start=1):
            row_to_use = ws.max_row + 1
            ws.cell(row=row_to_use, column=1).value = key
            ws.cell(row=row_to_use, column=2).value = val
    else:
        ws.cell(last_row, 1).value = data
    # Save the workbook
    wb.save(file_path)

# file_path = "C:\\Workspace\\Python\\UTH\\OrangeHRM\\test_output\\excel\\excel_2024-09-09_16-46-47\\exx.xlsx"
# sheet_name = "Sheet1"
# # Writing a list of data
# list_data = ["Item1", "Item2", "Item3"]
# write_data_to_excel(file_path, sheet_name, list_data)
#
# # Writing a single data item
# single_data = "SingleItem"
# write_data_to_excel(file_path, sheet_name, single_data)
