import openpyxl

def get_row_count(file, sheet_name):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(file, sheet_name):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_data_from_excel(file, sheet_name, row, column):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row, column).value

def write_data_into_excel(file, sheet_name, row, column, data):
    workbook = openpyxl.open(file)
    sheet = workbook[sheet_name]
    sheet.cell(row, column).value = data
    workbook.save(file)

def fetch_data_from_excel(file, sheet_name, row_header, column_header):
    try:
        workbook = openpyxl.open(file)
        sheet = workbook[sheet_name]
        row_index = None
        column_index = None
        for row in sheet.iter_rows(values_only=True):
            if row_header in row:
                row_index = row.index(row_header)
        for column in sheet.iter_cols(values_only=True):
            if column_header in column:
                column_index = column.index(column_header)

        if row_index is not None and column_index is not None:
            value = sheet.cell(row_index, column_index).value
            return value
        else:
            print(f"Headers not found: {row_header}, {column_header}")
            return None
    except Exception as e:
        print(f"Error loading workbook or fetching data: {e}")
        return None

